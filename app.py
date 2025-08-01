import streamlit as st
import requests
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

# Configuración de la página
st.set_page_config(page_title="Login SIGOF", layout="wide")

# Función para convertir fecha y hora
def convertir_fecha_hora(fecha_hora_str):
    meses = {
        "January": "01", "February": "02", "March": "03", "April": "04",
        "May": "05", "June": "06", "July": "07", "August": "08",
        "September": "09", "October": "10", "November": "11", "December": "12"
    }
    match = re.match(r"(\d{1,2}) de ([a-zA-Z]+) de (\d{4}) en horas: (\d{2}:\d{2}:\d{2})", fecha_hora_str)
    if match:
        dia, mes, anio, hora = match.groups()
        mes_num = meses.get(mes, "00")
        return f"{dia.zfill(2)}/{mes_num}/{anio} {hora}"
    return fecha_hora_str

# Interfaz de usuario
st.title("HUMANO INGRESE SUS CREDENCIALES DE SIGOF WEB")
usuario_input = st.text_input("👤 Usuario:")
password_input = st.text_input("🔑 Clave:", type="password")

if st.button("🔓 Humano inicia sesión"):
    login_url = "http://sigof.distriluz.com.pe/plus/usuario/login"
    data_url = "http://sigof.distriluz.com.pe/plus/ComlecOrdenlecturas/ajax_mostar_mapa_selfie"
    credentials = {
        "data[Usuario][usuario]": usuario_input,
        "data[Usuario][pass]": password_input
    }
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": login_url,
    }

    with requests.Session() as session:
        login_response = session.post(login_url, data=credentials, headers=headers)
        if "Usuario o contraseña incorrecto" in login_response.text:
            st.error("🧠 Humano, las credenciales son incorrectas.")
        else:
            data_response = session.get(data_url, headers=headers)
            data = data_response.text.replace("\\/", "/")
            data = re.sub(r"<\/?\w+.*?>", "", data)
            data = re.sub(r"\s+", " ", data).strip()
            blocks = re.split(r"Ver detalle", data)

            results = {}
            for block in blocks:
                fecha = re.search(r"Fecha Selfie:\s*(\d{1,2} de [a-zA-Z]+ de \d{4} en horas: \d{2}:\d{2}:\d{2})", block)
                lecturista = re.search(r"Lecturista:\s*([\w\sÁÉÍÓÚáéíóúÑñ]+)", block)
                url = re.search(r"url\":\"(https[^\"]+)", block)
                if fecha and lecturista and url:
                    fecha_hora_formateada = convertir_fecha_hora(fecha.group(1).strip())
                    fecha_selfie, _ = fecha_hora_formateada.split(" ")
                    lecturista_nombre = lecturista.group(1).strip()
                    url_imagen = url.group(1).strip()
                    key = (lecturista_nombre, fecha_selfie)
                    if key not in results:
                        results[key] = {"URLs Imagen": []}
                    results[key]["URLs Imagen"].append(url_imagen)

            if not results:
                st.warning("⚠️ Humano, tus credenciales son inválidas o no se encontraron datos.")
            else:
                max_urls = max(len(item["URLs Imagen"]) for item in results.values())
                url_columns = [f"Url_foto {i+1}" for i in range(max_urls)]
                vista_columns = [f"Vista Url_foto {i+1}" for i in range(max_urls)]
                columns = ["Fecha Selfie", "Lecturista"] + url_columns
                data = []
                for (lecturista, fecha_selfie), info in results.items():
                    row = [fecha_selfie, lecturista] + info["URLs Imagen"] + [""] * (max_urls - len(info["URLs Imagen"]))
                    data.append(row)

                df = pd.DataFrame(data, columns=columns)
                wb = Workbook()
                ws = wb.active
                ws.title = "LmcSelfiesLectura"
                ws.freeze_panes = "A2"
                ws.append(columns + vista_columns)

                # Estilo encabezado
                header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Ajuste de columnas
                ws.column_dimensions['A'].width = 11.5
                ws.column_dimensions['B'].width = 16
                for i in range(len(url_columns)):
                    col_letter = get_column_letter(3 + i)
                    ws.column_dimensions[col_letter].hidden = True

                for i, row in enumerate(df.itertuples(index=False), start=2):
                    row_data = list(row)
                    ws.append(row_data + [""] * max_urls)
                    ws[f'A{i}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    ws[f'B{i}'].alignment = Alignment(horizontal='justify', vertical='center', wrap_text=True)
                    for j in range(max_urls):
                        col_index = 3 + j
                        url_cell = f"{get_column_letter(col_index)}{i}"
                        vista_col_index = len(columns) + j + 1
                        formula_cell = f"{get_column_letter(vista_col_index)}{i}"
                        image_height_px = 200
                        image_width_px = 140
                        ws[formula_cell] = f'=IMAGE({url_cell},4,{image_height_px},{image_width_px})'
                        vista_col_letter = get_column_letter(vista_col_index)
                        ws.column_dimensions[vista_col_letter].width = round(image_width_px / 7, 1)
                    ws.row_dimensions[i].height = 151

                # Guardar el archivo localmente
                output_path = "Lmc_ReporteSelfie.xlsx"
                try:
                    wb.save(output_path)
                    st.success("✅ Humano, el archivo fue guardado correctamente.")
                    st.markdown(f'<a href="{output_path}" download><button>📥 Descargar archivo</button></a>', unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"Error al guardar el archivo: {e}")
